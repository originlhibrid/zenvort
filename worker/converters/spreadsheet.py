# worker/converters/spreadsheet.py
# Spreadsheet conversions via openpyxl + Python stdlib csv/json.
#
# Supported conversions:
#   xlsx → csv    (first sheet, comma-separated; sheet name via metadata)
#   xlsx → json   (array of row objects; first sheet; sheet name via metadata)
#   xlsx → html   (styled table, all sheets as tabs)
#   csv  → xlsx   (header row bold, auto column width)
#   json → xlsx   (array of objects, keys as column headers)
#
# Job metadata kwargs (passed by worker/tasks.py):
#   sheet_name  — for xlsx→csv/json: select a named sheet instead of first

import csv
import json
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from worker.security.path_guard import sanitize_and_assert_tmp_path

logger = logging.getLogger(__name__)


def convert(
    input_path: str,
    output_path: str,
    input_format: str,
    output_format: str,
    timeout_s: float = 120.0,
    *,
    sheet_name: str | None = None,
) -> None:
    sanitize_and_assert_tmp_path(input_path)

    if input_format == "xlsx":
        if output_format == "csv":
            _xlsx_to_csv(input_path, output_path, sheet_name=sheet_name)
        elif output_format == "json":
            _xlsx_to_json(input_path, output_path, sheet_name=sheet_name)
        elif output_format == "html":
            _xlsx_to_html(input_path, output_path)
        else:
            raise ValueError(
                f"spreadsheet does not support xlsx→{output_format}"
            )
    elif input_format == "csv":
        if output_format == "xlsx":
            _csv_to_xlsx(input_path, output_path)
        else:
            raise ValueError(
                f"spreadsheet does not support csv→{output_format}"
            )
    elif input_format == "json":
        if output_format == "xlsx":
            _json_to_xlsx(input_path, output_path)
        else:
            raise ValueError(
                f"spreadsheet does not support json→{output_format}"
            )
    else:
        raise ValueError(
            f"spreadsheet converter does not support {input_format}→{output_format}"
        )

    _assert_output(output_path, input_format, output_format)


def _assert_output(output_path: str, input_format: str, output_format: str) -> None:
    p = Path(output_path)
    if not p.exists() or p.stat().st_size == 0:
        raise RuntimeError(
            f"spreadsheet converter produced no output for "
            f"{input_format}→{output_format}"
        )


# ── xlsx → csv ───────────────────────────────────────────────────────────────

def _xlsx_to_csv(input_path: str, output_path: str, sheet_name: str | None = None) -> None:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    # Resolve target sheet.
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Sheet '{ws.title}' is empty")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            # Strip None → empty string, then write.
            writer.writerow(cell if cell is not None else "" for cell in row)

    logger.info(f"[spreadsheet] xlsx→csv ({ws.title}, {len(rows)} rows) → {output_path}")


# ── xlsx → json ──────────────────────────────────────────────────────────────

def _xlsx_to_json(input_path: str, output_path: str, sheet_name: str | None = None) -> None:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Sheet '{ws.title}' is empty")

    # First non-empty row → column headers.
    # Find the first row that has at least one non-None value.
    header_row_idx = None
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(f"Sheet '{ws.title}' has no data rows")

    headers = rows[header_row_idx]
    data_rows = rows[header_row_idx + 1:]

    objects: list[dict] = []
    for row in data_rows:
        obj = {
            str(headers[j]).strip() if headers[j] is not None else f"col{j + 1}":
                row[j] if j < len(row) and row[j] is not None else None
            for j in range(len(headers))
        }
        objects.append(obj)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(objects, f, indent=2, ensure_ascii=False, default=str)

    logger.info(f"[spreadsheet] xlsx→json ({ws.title}, {len(objects)} rows) → {output_path}")


# ── xlsx → html ──────────────────────────────────────────────────────────────

def _xlsx_to_html(input_path: str, output_path: str) -> None:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    # Header style: light blue bg + bold.
    th_fill   = PatternFill("solid", fgColor="D9E1F2")
    th_font   = Font(bold=True, color="1F3864")
    th_align  = Alignment(horizontal="left", vertical="center")
    td_align  = Alignment(horizontal="left", vertical="center")
    thin_side = Side(style="thin", color="B4C6E7")
    border    = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    lines: list[str] = []
    lines.append(
        "<!DOCTYPE html>\n"
        "<html lang='en'>\n<head>\n"
        "<meta charset='utf-8'/>\n"
        "<title>Spreadsheet</title>\n"
        "<style>\n"
        "  body { font-family: Arial, sans-serif; margin: 20px; }\n"
        "  .tab-bar { margin-bottom: 12px; }\n"
        "  .tab-btn  { padding: 6px 16px; margin-right: 4px; cursor: pointer; "
        "border: 1px solid #ccc; background: #f5f5f5; border-radius: 4px 4px 0 0; }\n"
        "  .tab-btn.active { background: #fff; border-bottom: 2px solid #fff; font-weight: bold; }\n"
        "  .tab-content { display: none; }\n"
        "  .tab-content.active { display: block; }\n"
        "  table { border-collapse: collapse; font-size: 13px; }\n"
        "  th { background: #D9E1F2; font-weight: bold; color: #1F3864; "
        "text-align: left; padding: 6px 10px; border: 1px solid #B4C6E7; white-space: nowrap; }\n"
        "  td { padding: 5px 10px; border: 1px solid #B4C6E7; "
        "vertical-align: top; max-width: 400px; overflow: hidden; text-overflow: ellipsis; }\n"
        "  tr:nth-child(even) td { background: #F5F8FF; }\n"
        "  .col-num { color: #888; font-size: 11px; text-align: right; }\n"
        "</style>\n"
        "</head>\n<body>\n"
        "<h2>Spreadsheet Export</h2>\n"
    )

    sheet_names = wb.sheetnames
    if not sheet_names:
        wb.close()
        raise ValueError("Workbook has no sheets")

    # Tab bar.
    if len(sheet_names) > 1:
        lines.append("<div class='tab-bar'>\n")
        for i, name in enumerate(sheet_names):
            active = " active" if i == 0 else ""
            safe_name = name.replace("'", "\\'")
            lines.append(
                f"  <button class='tab-btn{active}' "
                f"onclick=\"document.querySelectorAll('.tab-content').forEach(c=>c.classList.remove('active'));"
                f"document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));"
                f"document.getElementById('sheet-{i}').classList.add('active');"
                f"this.classList.add('active');\">{safe_name}</button>\n"
            )
        lines.append("</div>\n")

    for sheet_idx, sheet_name in enumerate(sheet_names):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        tab_active = "" if sheet_idx > 0 else " active"
        lines.append(f"<div id='sheet-{sheet_idx}' class='tab-content{tab_active}'>\n")
        lines.append(f"<h3>{sheet_name}</h3>\n")
        lines.append("<table>\n")

        for r_idx, row in enumerate(rows):
            # Skip entirely-empty rows.
            if all(cell is None for cell in row):
                continue
            tag = "th" if r_idx == 0 else "td"
            lines.append("<tr>")
            for cell in row:
                val = cell if cell is not None else ""
                # Escape HTML special chars.
                val_str = str(val).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                style = ""
                if r_idx == 0:
                    style = "background:#D9E1F2;color:#1F3864;font-weight:bold"
                elif r_idx % 2 == 0:
                    style = "background:#F5F8FF"
                style_attr = f" style='{style}'" if style else ""
                lines.append(f"<{tag}{style_attr}>{val_str}</{tag}>")
            lines.append("</tr>\n")
        lines.append("</table>\n</div>\n")

    lines.append(
        "<script>\n"
        "  document.querySelectorAll('.tab-btn').forEach(function(btn) {\n"
        "    btn.addEventListener('click', function() {\n"
        "      document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));\n"
        "      document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));\n"
        "      var sheetId = btn.textContent.trim().replace(/[^a-zA-Z0-9_-]/g,'_');\n"
        "      var target = document.getElementById('sheet-' + Array.from(document.querySelectorAll('.tab-btn')).indexOf(btn));\n"
        "      if (target) { target.classList.add('active'); btn.classList.add('active'); }\n"
        "    });\n"
        "  });\n"
        "</script>\n"
    )
    lines.append("</body>\n</html>")

    Path(output_path).write_text("".join(lines), encoding="utf-8")
    logger.info(
        f"[spreadsheet] xlsx→html ({len(sheet_names)} sheets, "
        f"{sum(1 for s in sheet_names for _ in wb[s].iter_rows())} total rows) → {output_path}"
    )
    wb.close()


# ── csv → xlsx ───────────────────────────────────────────────────────────────

def _csv_to_xlsx(input_path: str, output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    with open(input_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Bold headers + auto column width.
    _style_xlsx_sheet(ws, has_header=True)
    wb.save(output_path)
    logger.info(f"[spreadsheet] csv→xlsx → {output_path}")


# ── json → xlsx ──────────────────────────────────────────────────────────────

def _json_to_xlsx(input_path: str, output_path: str) -> None:
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects")

    if not data:
        raise ValueError("JSON array is empty")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Collect all keys across all objects for column order.
    all_keys: list[str] = []
    for obj in data:
        if isinstance(obj, dict):
            for k in obj:
                if k not in all_keys:
                    all_keys.append(k)

    if not all_keys:
        raise ValueError("JSON objects have no keys")

    # Write header row.
    ws.append(all_keys)

    # Write data rows.
    for obj in data:
        row = [obj.get(k) for k in all_keys]
        ws.append(row)

    _style_xlsx_sheet(ws, has_header=True)
    wb.save(output_path)
    logger.info(f"[spreadsheet] json→xlsx ({len(data)} rows, {len(all_keys)} cols) → {output_path}")


def _style_xlsx_sheet(ws, has_header: bool) -> None:
    """Apply bold headers + auto column width + cell padding to a worksheet."""
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True, color="1F3864")
    center      = Alignment(horizontal="left", vertical="center", wrap_text=False)

    thin_side = Side(style="thin", color="B4C6E7")
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    # Header row.
    if has_header:
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border

    # Auto width per column.
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len    = 0
        for cell in col_cells:
            try:
                cell_val = str(cell.value) if cell.value is not None else ""
                # Account for wrapping — use 60% of actual length estimate.
                max_len = max(max_len, min(len(cell_val), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # Cell padding + border for all cells.
    for row in ws.iter_rows():
        for cell in row:
            if not cell.alignment or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border